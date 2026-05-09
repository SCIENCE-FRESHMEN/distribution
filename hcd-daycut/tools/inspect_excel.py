import sys

import openpyxl


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def main(argv: list[str]) -> int:
    path = argv[1] if len(argv) > 1 else "库管系统算法升级接口测试记录_20260419.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print("sheets:", wb.sheetnames)

    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n== {sname} == rows={ws.max_row} cols={ws.max_column}")
        max_row = min(ws.max_row or 0, 60)
        max_col = min(ws.max_column or 0, 40)
        for r in range(1, max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            if all(_is_blank(v) for v in row):
                continue
            print(r, row)

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

