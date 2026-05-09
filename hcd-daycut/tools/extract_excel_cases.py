import json
import sys
from dataclasses import dataclass
from typing import Any

import openpyxl


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


@dataclass(frozen=True)
class Case:
    sheet: str
    row: int
    seq: str
    api: str
    ok_ng: str
    close_status: str
    target: str
    problem: str
    request: str
    expected_response: str


def _get(ws, r: int, c: int) -> Any:
    return ws.cell(r, c).value


def _find_col(headers: dict[str, int], *names: str) -> int:
    for n in names:
        if n in headers:
            return headers[n]
    return 0


def extract(path: str) -> list[Case]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    cases: list[Case] = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row is None or ws.max_row < 2:
            continue

        # Header on row 1
        headers: dict[str, int] = {}
        for c in range(1, (ws.max_column or 1) + 1):
            hv = _norm_header(_get(ws, 1, c))
            if hv:
                headers[hv] = c

        col_seq = _find_col(headers, "序号")
        col_api = _find_col(headers, "接口")
        col_target = _find_col(headers, "测试目标")
        col_req = _find_col(headers, "请求报文")
        col_resp = _find_col(headers, "响应报文")
        col_ok = _find_col(headers, "OK/NG")
        col_close = _find_col(headers, "关闭状态")
        col_problem = _find_col(headers, "问题点", "说明")

        if not col_seq and not col_api:
            continue

        for r in range(2, (ws.max_row or 2) + 1):
            seq = _get(ws, r, col_seq) if col_seq else None
            api = _get(ws, r, col_api) if col_api else None
            if _is_blank(seq) and _is_blank(api):
                # Many sheets have trailing empty rows.
                continue

            case = Case(
                sheet=sname,
                row=r,
                seq="" if seq is None else str(seq).strip(),
                api="" if api is None else str(api).strip(),
                ok_ng="" if not col_ok or _get(ws, r, col_ok) is None else str(_get(ws, r, col_ok)).strip(),
                close_status=""
                if not col_close or _get(ws, r, col_close) is None
                else str(_get(ws, r, col_close)).strip(),
                target=""
                if not col_target or _get(ws, r, col_target) is None
                else str(_get(ws, r, col_target)).strip(),
                problem=""
                if not col_problem or _get(ws, r, col_problem) is None
                else str(_get(ws, r, col_problem)).strip(),
                request=""
                if not col_req or _get(ws, r, col_req) is None
                else str(_get(ws, r, col_req)).strip(),
                expected_response=""
                if not col_resp or _get(ws, r, col_resp) is None
                else str(_get(ws, r, col_resp)).strip(),
            )
            cases.append(case)

    return cases


def main(argv: list[str]) -> int:
    default_path = "库管系统算法升级接口测试记录_20260419.xlsx"
    path = default_path
    for arg in argv[1:]:
        if arg.startswith("-"):
            continue
        path = arg
        break
    cases = extract(path)

    open_ng = [
        c
        for c in cases
        if c.ok_ng in ("❌", "NG", "ng")
        and c.close_status not in ("已关闭", "关闭", "Closed")
    ]
    all_ng = [c for c in cases if c.ok_ng in ("❌", "NG", "ng")]

    print("total_cases:", len(cases))
    print("open_ng_cases:", len(open_ng))
    for c in open_ng:
        print(f"- [{c.sheet}] seq={c.seq} api={c.api} row={c.row}")
        if c.target:
            print("  target:", c.target.replace("\n", " ")[:160])
        if c.problem:
            print("  problem:", c.problem.replace("\n", " ")[:200])

    if "--all-ng" in argv:
        print("\nALL_NG_BEGIN")
        for c in all_ng:
            print(f"- [{c.sheet}] seq={c.seq} api={c.api} ok_ng={c.ok_ng} close={c.close_status} row={c.row}")
            if c.problem:
                print("  problem:", c.problem.replace("\n", " ")[:240])
        print("ALL_NG_END")

    # Also emit JSON to stdout if requested
    if "--json" in argv:
        payload = [c.__dict__ for c in cases]
        print("\nJSON_BEGIN")
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        print("JSON_END")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
